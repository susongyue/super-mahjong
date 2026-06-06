"""
角色数据双向同步脚本
=====================
xlsx <-> JSON 互转，按 B 列角色 ID（SRM####）匹配。

格式保护：
- 写入时只修改单元格 .value，不触碰字体/颜色/边框/对齐等格式
- 首次从 xlsx 读取时自动提取格式模板（从第一行数据行）
- 对新写入值的空单元格自动应用模板格式，保持一致性
- 保留原 xlsx 的行高、列宽、冻结窗格、条件格式等所有结构

用法:
    python tools/sync_characters.py --to-json          # xlsx → json
    python tools/sync_characters.py --to-xlsx          # json → xlsx
    python tools/sync_characters.py --to-json --xlsx path/to/file.xlsx
    python tools/sync_characters.py --to-json --json path/to/file.json
"""

import copy
import json
import argparse
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    sys.exit(1)

# 默认路径
PROJECT_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = PROJECT_ROOT / "super-mahjong.xlsx"
DEFAULT_JSON = PROJECT_ROOT / "data" / "characters.json"
SHEET_NAME = "修改版"

def _get_sheet(wb):
    """获取工作表：优先用 SHEET_NAME，找不到则用第一个 sheet"""
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    if wb.sheetnames:
        print(f"[WARN] 未找到 '{SHEET_NAME}'，改用 '{wb.sheetnames[0]}'")
        return wb[wb.sheetnames[0]]
    raise KeyError("xlsx 文件中没有任何工作表")

# 列映射（1-based）
COL_FACTION = 1   # A: 阵营
COL_ID      = 2   # B: 角色ID
COL_NAME    = 3   # C: 角色名
COL_TIER    = 4   # D: 梯度
COL_LIMIT   = 5   # E: 限定技
COL_PASSIVE = 6   # F: 被动技
COL_INNATE  = 7   # G: 固有技
COL_EXTRA   = 8   # H: 附加技
COL_NOTE    = 9   # I: 说明
COL_EN      = 10  # J: 英文名
COL_KANA    = 11  # K: 片假名


def read_xlsx(xlsx_path: Path) -> list[dict]:
    """从 xlsx 读取角色数据，以列表返回"""
    wb = load_workbook(str(xlsx_path))
    ws = _get_sheet(wb)

    chars = []
    for row in range(2, ws.max_row + 1):
        cid = ws.cell(row=row, column=COL_ID).value
        if cid is None:
            continue

        en = ws.cell(row=row, column=COL_EN).value or ""
        image = f"{en}.png" if en else ""

        c = {
            "id": cid,
            "faction": ws.cell(row=row, column=COL_FACTION).value or "",
            "name": ws.cell(row=row, column=COL_NAME).value or "",
            "tier": ws.cell(row=row, column=COL_TIER).value or "",
            "limit": ws.cell(row=row, column=COL_LIMIT).value or "",
            "passive": ws.cell(row=row, column=COL_PASSIVE).value or "",
            "innate": ws.cell(row=row, column=COL_INNATE).value or "",
            "extra": ws.cell(row=row, column=COL_EXTRA).value or "",
            "note": ws.cell(row=row, column=COL_NOTE).value or "",
            "image": image,
        }
        chars.append(c)

    wb.close()
    return chars


def write_json(chars: list[dict], json_path: Path):
    """将角色列表写入 JSON 文件"""
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(chars, f, ensure_ascii=False, indent=2)
    print(f"[OK] 已写入 {len(chars)} 个角色到 {json_path}")


def read_json(json_path: Path) -> list[dict]:
    """从 JSON 读取角色数据"""
    with open(json_path, "r", encoding="utf-8") as f:
        return json.load(f)


def xlsx_to_json(xlsx_path: Path, json_path: Path):
    """xlsx → json：按行读取 xlsx，全量覆盖 json"""
    chars = read_xlsx(xlsx_path)
    write_json(chars, json_path)


def _extract_template_style(ws, row) -> dict:
    """
    从 xlsx 工作表的一行提取格式模板。
    返回 {column_index: openpyxl cell} 的格式快照字典。
    优先使用有内容的单元格作为模板。
    """
    template = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if cell.value is not None:
            template[col] = {
                "font": copy.copy(cell.font),
                "fill": copy.copy(cell.fill),
                "border": copy.copy(cell.border),
                "alignment": copy.copy(cell.alignment),
                "number_format": cell.number_format,
            }
    return template


def _ensure_cell_format(cell, template: dict, col: int):
    """
    确保 cell 的格式与模板一致。
    对 openpyxl 来说，只设 value 不改变已有格式，
    但为了防御性编程，确保写入的单元格都有正确格式。
    优先保留已有格式，只在完全没有格式时应用模板。
    """
    if col not in template:
        return
    fmt = template[col]
    # 只在单元格没有自定义格式时应用模板
    # (openpyxl 中 cell.font.name 默认是 'Calibri' 如果从未设置过)
    if cell.font.name == "Calibri" and cell.font.size == 11.0:
        cell.font = copy.copy(fmt["font"])
        cell.fill = copy.copy(fmt["fill"])
        cell.border = copy.copy(fmt["border"])
        cell.alignment = copy.copy(fmt["alignment"])
        cell.number_format = fmt["number_format"]


def json_to_xlsx(json_path: Path, xlsx_path: Path):
    """
    json → xlsx：按 id 匹配更新 xlsx 对应行。

    格式保护策略：
    1. 先加载 xlsx，从第 2 行提取数据行格式模板
    2. 只更新单元格的 .value，不修改任何格式属性
    3. 对之前为空的单元格，应用模板格式来保持一致性
    4. 保留原 xlsx 的行高、列宽、冻结窗格等所有结构
    """
    chars = read_json(json_path)
    json_by_id = {c["id"]: c for c in chars}

    wb = load_workbook(str(xlsx_path))
    ws = _get_sheet(wb)

    # ---- 1. 提取格式模板 ----
    # 找第一个有内容的数据行作为格式模板
    template_row = 2
    for r in range(2, ws.max_row + 1):
        cid = ws.cell(row=r, column=COL_ID).value
        if cid is not None:
            template_row = r
            break
    data_template = _extract_template_style(ws, template_row)
    header_template = _extract_template_style(ws, 1)

    # 合并模板：优先用 data_template，缺的用 header_template
    full_template = {}
    for col in range(1, ws.max_column + 1):
        full_template[col] = data_template.get(col) or header_template.get(col)

    # ---- 2. 建立 xlsx id→row 映射 ----
    xlsx_id_to_row = {}
    for row in range(2, ws.max_row + 1):
        xlsx_id = ws.cell(row=row, column=COL_ID).value
        if xlsx_id is not None:
            xlsx_id_to_row[xlsx_id] = row

    # ---- 3. 逐行更新 ----
    updated = 0
    not_found = []
    skipped = []

    for row in range(2, ws.max_row + 1):
        xlsx_id = ws.cell(row=row, column=COL_ID).value
        if xlsx_id is None:
            continue

        c = json_by_id.get(xlsx_id)
        if c is None:
            not_found.append(xlsx_id)
            continue

        xlsx_name = ws.cell(row=row, column=COL_NAME).value
        if c.get("name") != xlsx_name:
            skipped.append(f"  {xlsx_id}: xlsx={xlsx_name}  json={c.get('name')}")
            continue

        # 只更新 .value，不碰格式
        # 注：COL_KANA(片假名) 不在 JSON schema 中，不覆盖，保留 xlsx 原值
        col_values = {
            COL_FACTION: c.get("faction") or None,
            COL_NAME: c.get("name") or None,
            COL_TIER: c.get("tier") or None,
            COL_LIMIT: c.get("limit") or None,
            COL_PASSIVE: c.get("passive") or None,
            COL_INNATE: c.get("innate") or None,
            COL_EXTRA: c.get("extra") or None,
            COL_NOTE: c.get("note") or None,
        }

        img = c.get("image", "")
        en = img.replace(".png", "") if img else None
        col_values[COL_EN] = en

        for col, val in col_values.items():
            cell = ws.cell(row=row, column=col)
            cell.value = val
            _ensure_cell_format(cell, full_template, col)

        updated += 1

    # ---- 4. 处理 JSON 有但 xlsx 没有的新角色 ----
    json_only = set(json_by_id.keys()) - set(xlsx_id_to_row.keys())

    if json_only:
        new_rows_added = 0
        for rid in sorted(json_only):
            c = json_by_id[rid]
            row = ws.max_row + 1 + new_rows_added

            img = c.get("image", "")
            en = img.replace(".png", "") if img else None

            col_values = {
                COL_FACTION: c.get("faction") or None,
                COL_ID: rid,
                COL_NAME: c.get("name") or None,
                COL_TIER: c.get("tier") or None,
                COL_LIMIT: c.get("limit") or None,
                COL_PASSIVE: c.get("passive") or None,
                COL_INNATE: c.get("innate") or None,
                COL_EXTRA: c.get("extra") or None,
                COL_NOTE: c.get("note") or None,
                COL_EN: en,
            }

            for col, val in col_values.items():
                cell = ws.cell(row=row, column=col)
                cell.value = val
                _ensure_cell_format(cell, full_template, col)

            new_rows_added += 1

        if new_rows_added > 0:
            print(f"[INFO] 新增 {new_rows_added} 个角色到 xlsx 末尾")

    # ---- 5. 保存 ----
    wb.save(str(xlsx_path))
    wb.close()

    print(f"[OK] 已更新 {updated} 个角色到 {xlsx_path}")

    if not_found:
        print(f"\n[WARN] xlsx 中有 {len(not_found)} 个 ID 在 json 中未找到 (保留原样):")
        for rid in not_found:
            print(f"  {rid}")

    if skipped:
        print(f"\n[WARN] {len(skipped)} 个角色名称不匹配 (未更新):")
        for s in skipped:
            print(s)

    if json_only:
        print(f"\n[INFO] json 中有 {len(json_only)} 个 ID 在 xlsx 中原本不存在 (已追加):")
        for rid in sorted(json_only):
            print(f"  {rid}: {json_by_id[rid].get('name')}")


def main():
    parser = argparse.ArgumentParser(description="角色数据 xlsx <-> json 双向同步")
    group = parser.add_mutually_exclusive_group(required=True)
    group.add_argument("--to-json", action="store_true", help="xlsx → json")
    group.add_argument("--to-xlsx", action="store_true", help="json → xlsx")

    parser.add_argument("--xlsx", type=Path, default=DEFAULT_XLSX,
                        help=f"xlsx 文件路径 (默认: {DEFAULT_XLSX})")
    parser.add_argument("--json", type=Path, default=DEFAULT_JSON,
                        help=f"json 文件路径 (默认: {DEFAULT_JSON})")

    args = parser.parse_args()

    if not args.xlsx.exists():
        print(f"[ERROR] xlsx 文件不存在: {args.xlsx}")
        sys.exit(1)

    if args.to_json:
        print(f"xlsx → json")
        print(f"  源: {args.xlsx}")
        print(f"  目标: {args.json}")
        xlsx_to_json(args.xlsx, args.json)
    else:
        if not args.json.exists():
            print(f"[ERROR] json 文件不存在: {args.json}")
            sys.exit(1)
        print(f"json → xlsx")
        print(f"  源: {args.json}")
        print(f"  目标: {args.xlsx}")
        json_to_xlsx(args.json, args.xlsx)


if __name__ == "__main__":
    main()
